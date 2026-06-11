"""Excel writer — produces the output .xlsx with color-coded rows.

KEEP = green row, REVIEW = yellow row, DROP_* = red row.
Uses context managers and explicit .close() to prevent FD leaks.
All writes use str() coercion to avoid pandas 3 dtype crashes.
"""
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Row colors
_GREEN = PatternFill("solid", fgColor="D4EDDA")
_YELLOW = PatternFill("solid", fgColor="FFF3CD")
_RED = PatternFill("solid", fgColor="F8D7DA")
_GREY = PatternFill("solid", fgColor="F8F9FA")
_HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_BODY_FONT = Font(name="Arial", size=9)
_LINK_FONT = Font(name="Arial", size=9, color="4DA8FF", underline="single")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Columns we produce (in order). Original CSV columns are preserved too.
# Social columns (Instagram/YouTube/Facebook) removed — socials live in Genitractor.
AUDIT_COLUMNS = [
    "Status",
    "Status Reason",
    "iTunes Labels",
    "Deezer Labels",
    "Earliest Year",
    "AI Note",
]

# Preferred column widths
_WIDTHS = {
    "Artist": 22, "Spotify Links": 42, "Genres": 26, "Region": 14,
    "Spotify Monthly Listeners": 16, "Associated Labels": 22,
    "Recent Momentum": 13, "Status": 14, "Status Reason": 52,
    "iTunes Labels": 40, "Deezer Labels": 32,
    "Earliest Year": 12, "AI Note": 30,
}


def write_xlsx(df: pd.DataFrame, path: Path):
    """Write the audit dataframe to an xlsx file with formatting.

    Uses explicit close() to prevent FD leaks on macOS.
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Virtual Scout"

    # Determine columns to write: original CSV columns + audit columns
    headers = list(df.columns)

    # Header row
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=ci)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[1].height = 26

    # Data rows
    for ri, row_tuple in enumerate(df.itertuples(index=False), 2):
        row_dict = dict(zip(headers, row_tuple))
        status = str(row_dict.get("Status", ""))

        # Determine row fill
        if status == "KEEP":
            row_fill = _GREEN
        elif status == "REVIEW":
            row_fill = _YELLOW
        elif status.startswith("DROP"):
            row_fill = _RED
        elif ri % 2 == 0:
            row_fill = _GREY
        else:
            row_fill = None

        for ci, val in enumerate(row_tuple, 1):
            cell = ws.cell(row=ri, column=ci, value=str(val) if val is not None and str(val) != "nan" else "")
            cell.font = _BODY_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(headers[ci-1] in ("Status Reason", "iTunes Labels")))
            if row_fill:
                cell.fill = row_fill

        # Hyperlink Spotify column
        if "Spotify Links" in headers:
            sp_idx = headers.index("Spotify Links") + 1
            sp_val = str(row_dict.get("Spotify Links", ""))
            if sp_val.startswith("http"):
                cell = ws.cell(row=ri, column=sp_idx)
                cell.hyperlink = sp_val
                cell.font = _LINK_FONT

    # Column widths
    for ci, col_name in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _WIDTHS.get(col_name, 16)

    # Freeze header + autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Save and explicitly close
    try:
        wb.save(str(path))
    finally:
        wb.close()


def filter_xlsx_by_status(source_path: Path, dest_path: Path, statuses: List[str]) -> int:
    """Read an existing xlsx, filter rows by Status column, write a new xlsx.

    Args:
        source_path: Path to the full output xlsx
        dest_path: Path to write the filtered xlsx
        statuses: List of status values to include (e.g. ["KEEP"])
                  Special value "ALL" includes everything.
                  Special value "DROP" includes all DROP_* variants.

    Returns:
        Number of rows kept in the filtered output.

    Uses explicit close() to prevent FD leaks.
    """
    dest_path.parent.mkdir(parents=True, exist_ok=True)

    wb_src = load_workbook(str(source_path), read_only=True)
    try:
        ws_src = wb_src.active
        rows = list(ws_src.iter_rows(values_only=True))
    finally:
        wb_src.close()

    if not rows:
        return 0

    headers = list(rows[0])
    data_rows = rows[1:]

    # Find Status column
    try:
        status_idx = headers.index("Status")
    except ValueError:
        # No Status column — write all rows
        status_idx = None

    # Filter
    if "ALL" in statuses:
        filtered = data_rows
    elif status_idx is not None:
        filtered = []
        for row in data_rows:
            row_status = str(row[status_idx]) if row[status_idx] else ""
            if row_status in statuses:
                filtered.append(row)
            elif "DROP" in statuses and row_status.startswith("DROP"):
                filtered.append(row)
    else:
        filtered = data_rows

    if not filtered:
        return 0

    # Rebuild as a DataFrame and write
    df = pd.DataFrame(filtered, columns=headers)
    write_xlsx(df, dest_path)
    return len(filtered)



def merge_all_outputs(source_paths: list, dest_path: Path) -> Path:
    """Merge multiple output xlsx files into one combined sheet.

    Reads all source files, concatenates their data rows (preserving headers
    from the first file), and writes a single output with formatting.
    Uses explicit close() on all workbooks to prevent FD leaks.
    """
    dest_path.parent.mkdir(parents=True, exist_ok=True)

    all_rows = []
    headers = None

    for src_path in source_paths:
        wb = load_workbook(str(src_path), read_only=True)
        try:
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if not rows:
            continue

        if headers is None:
            headers = list(rows[0])

        # Add data rows (skip header row of each file)
        for row in rows[1:]:
            all_rows.append(row)

    if not headers or not all_rows:
        raise ValueError("No data rows found in any output file")

    # Build DataFrame and write
    df = pd.DataFrame(all_rows, columns=headers)
    write_xlsx(df, dest_path)
    return dest_path
