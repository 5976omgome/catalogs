"""Flask server — routes for upload, queue control, SSE stream, export, settings.

SSE note: does NOT set 'Connection' header (that was the hop-by-hop crash
on Waitress). Waitress handles keep-alive/close itself.
"""
import json
import time
import uuid
from pathlib import Path

from flask import Flask, request, Response, jsonify, send_file

from app import config, excel
from app.jobs import JobManager

app = Flask(__name__, static_folder="static", static_url_path="")
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024  # 32 MB max upload

_manager = JobManager()


def get_manager() -> JobManager:
    return _manager


# ---------------------------------------------------------------------------
# Static / index
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return app.send_static_file("index.html")


@app.route("/genitractor")
def genitractor_page():
    return app.send_static_file("genitractor.html")


# ---------------------------------------------------------------------------
# Settings (API keys)
# ---------------------------------------------------------------------------

@app.route("/api/status")
def api_status():
    store = config.keys_store()
    s = store.status()
    return jsonify({
        "groq_set": s["groq_api_key"]["set"],
        "gemini_set": s["gemini_api_key"]["set"],
        "genius_set": s["genius_token"]["set"],
    })


@app.route("/api/settings", methods=["GET"])
def api_settings_get():
    store = config.keys_store()
    return jsonify(store.status())


@app.route("/api/settings", methods=["POST"])
def api_settings_post():
    data = request.get_json(silent=True)
    if not data or not isinstance(data, dict):
        return jsonify({"error": "invalid body"}), 400
    store = config.keys_store()
    for key in ("groq_api_key", "gemini_api_key", "genius_token"):
        if key in data:
            val = str(data[key]).strip()
            if val:
                store.set(key, val)
            else:
                store.clear(key)
    return jsonify(store.status())


@app.route("/api/settings/clear", methods=["POST"])
def api_settings_clear():
    store = config.keys_store()
    store.clear()
    return jsonify(store.status())


# ---------------------------------------------------------------------------
# File upload
# ---------------------------------------------------------------------------

@app.route("/api/upload", methods=["POST"])
def api_upload():
    if "file" not in request.files:
        return jsonify({"error": "no file"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "no filename"}), 400

    # Validate extension
    ext = Path(f.filename).suffix.lower()
    if ext not in (".csv", ".tsv"):
        return jsonify({"error": "only .csv and .tsv files accepted"}), 400

    # Save with UUID prefix to avoid collisions
    safe_name = f"{uuid.uuid4().hex[:8]}_{Path(f.filename).name}"
    dest = config.UPLOAD_DIR / safe_name
    f.save(str(dest))

    mgr = get_manager()
    item = mgr.add(filename=f.filename, path=dest)
    return jsonify(item.to_dict()), 201


# ---------------------------------------------------------------------------
# Queue control
# ---------------------------------------------------------------------------

@app.route("/api/queue/start", methods=["POST"])
def api_queue_start():
    mgr = get_manager()
    mgr.start()
    return jsonify({"ok": True})


@app.route("/api/queue/stop", methods=["POST"])
def api_queue_stop():
    mgr = get_manager()
    mgr.stop()
    return jsonify({"ok": True})


@app.route("/api/queue/clear", methods=["POST"])
def api_queue_clear():
    mgr = get_manager()
    mgr.clear_done()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Download / Export
# ---------------------------------------------------------------------------

@app.route("/api/download/<item_id>")
def api_download(item_id: str):
    """Download the full output xlsx (works on running/stopped/error items too)."""
    mgr = get_manager()
    item = mgr.find(item_id)
    if not item:
        return jsonify({"error": "not found"}), 404
    if not item.output_path or not item.output_path.exists():
        return jsonify({"error": "no output yet — run not started or no artists processed"}), 404
    return send_file(
        str(item.output_path),
        as_attachment=True,
        download_name=item.output_path.name,
    )


@app.route("/api/export/<item_id>/<filter_name>")
def api_export(item_id: str, filter_name: str):
    """Export a filtered subset of the output xlsx."""
    mgr = get_manager()
    item = mgr.find(item_id)
    if not item:
        return jsonify({"error": "not found"}), 404
    if not item.output_path or not item.output_path.exists():
        return jsonify({"error": "no output yet"}), 404

    f = filter_name.lower()
    if f == "keep":
        statuses = ["KEEP"]
    elif f == "review":
        statuses = ["REVIEW"]
    elif f == "drops":
        statuses = ["DROP"]
    elif f == "all":
        statuses = ["ALL"]
    else:
        return jsonify({"error": "filter must be keep|review|drops|all"}), 400

    stem = item.output_path.stem
    dst = config.OUTPUT_DIR / f"{stem}-{f}.xlsx"
    try:
        kept = excel.filter_xlsx_by_status(item.output_path, dst, statuses)
    except Exception as e:
        return jsonify({"error": f"export failed: {e}"}), 500

    if kept == 0:
        try:
            dst.unlink(missing_ok=True)
        except Exception:
            pass
        return jsonify({"error": "no rows match this filter"}), 404

    return send_file(str(dst), as_attachment=True, download_name=dst.name)


@app.route("/api/export_all")
def api_export_all():
    """Merge ALL finished item outputs into one master xlsx."""
    mgr = get_manager()
    items_with_output = []
    with mgr._lock:
        for item in mgr._items:
            if item.output_path and item.output_path.exists():
                items_with_output.append(item)

    if not items_with_output:
        return jsonify({"error": "no finished outputs to merge"}), 404

    try:
        merged_path = excel.merge_all_outputs(
            [item.output_path for item in items_with_output],
            config.OUTPUT_DIR / "AllCombinedOutput.xlsx",
        )
    except Exception as e:
        return jsonify({"error": f"merge failed: {e}"}), 500

    return send_file(
        str(merged_path),
        as_attachment=True,
        download_name="AllCombinedOutput.xlsx",
    )


@app.route("/api/queue/stop_and_export/<item_id>")
def api_stop_and_export(item_id: str):
    """Safety net: stop the run AND download whatever's been processed."""
    mgr = get_manager()
    item = mgr.find(item_id)
    if not item:
        return jsonify({"error": "not found"}), 404

    mgr.stop()

    # Give worker up to 2s to flush final checkpoint
    for _ in range(20):
        if item.output_path and item.output_path.exists():
            break
        time.sleep(0.1)

    if not item.output_path or not item.output_path.exists():
        return jsonify({"error": "no output yet"}), 404

    return send_file(
        str(item.output_path),
        as_attachment=True,
        download_name=item.output_path.name,
    )


# ---------------------------------------------------------------------------
# Genius Social Pass — runs AFTER main audit, one artist per 2 seconds
# ---------------------------------------------------------------------------

_genius_thread = None
_genius_running = False
_genius_stop = False


@app.route("/api/genius/run", methods=["POST"])
def api_genius_run():
    """Start a slow Genius social pass on all finished items."""
    global _genius_thread, _genius_running, _genius_stop

    if _genius_running:
        return jsonify({"error": "Genius pass already running"}), 409

    key = config.genius_token()
    if not key:
        return jsonify({"error": "No Genius token configured"}), 400

    _genius_stop = False

    import threading
    _genius_thread = threading.Thread(target=_genius_worker, daemon=True)
    _genius_thread.start()

    return jsonify({"ok": True, "message": "Genius pass started"})


@app.route("/api/genius/stop", methods=["POST"])
def api_genius_stop():
    """Stop the running Genius pass."""
    global _genius_stop
    _genius_stop = True
    return jsonify({"ok": True})


@app.route("/api/genius/status")
def api_genius_status():
    """Check if Genius pass is running."""
    return jsonify({"running": _genius_running})


def _genius_worker():
    """Process all finished items, updating xlsx with socials. 1 artist per 2 sec."""
    global _genius_running, _genius_stop
    _genius_running = True

    from app.sources import genius
    import openpyxl
    import time as _time

    mgr = get_manager()

    try:
        with mgr._lock:
            items = [i for i in mgr._items if i.output_path and i.output_path.exists()]

        if not items:
            print("[genius-pass] No finished items to process.", flush=True)
            return

        total_found = 0
        total_processed = 0

        for item in items:
            if _genius_stop:
                print("[genius-pass] Stopped by user.", flush=True)
                break

            try:
                wb = openpyxl.load_workbook(str(item.output_path))
                ws = wb.active
            except Exception as e:
                print(f"[genius-pass] Failed to open {item.output_path.name}: {e}", flush=True)
                continue

            # Find column indices
            headers = {cell.value: cell.column for cell in ws[1] if cell.value}
            artist_col = headers.get("artist") or headers.get("Artist") or headers.get("artist name")
            ig_col = headers.get("Instagram")
            fb_col = headers.get("Facebook")

            if not artist_col:
                print(f"[genius-pass] No artist column in {item.output_path.name}", flush=True)
                continue

            # If Instagram/Facebook columns don't exist, add them
            if not ig_col:
                ig_col = ws.max_column + 1
                ws.cell(row=1, column=ig_col, value="Instagram")
            if not fb_col:
                fb_col = ws.max_column + 1
                ws.cell(row=1, column=fb_col, value="Facebook")

            modified = False

            for row_idx in range(2, ws.max_row + 1):
                if _genius_stop:
                    break

                artist_name = ws.cell(row=row_idx, column=artist_col).value
                if not artist_name:
                    continue

                # Skip if already has socials
                existing_ig = ws.cell(row=row_idx, column=ig_col).value
                existing_fb = ws.cell(row=row_idx, column=fb_col).value
                if existing_ig or existing_fb:
                    continue

                total_processed += 1

                # Rate limit: 1 request per 2 seconds
                _time.sleep(2.0)

                socials = genius.get_socials(str(artist_name).strip())

                if socials:
                    if socials.get("instagram"):
                        ws.cell(row=row_idx, column=ig_col, value=f"https://instagram.com/{socials['instagram']}")
                        modified = True
                    if socials.get("facebook"):
                        fb = socials["facebook"]
                        ws.cell(row=row_idx, column=fb_col, value=fb if fb.startswith("http") else f"https://facebook.com/{fb}")
                        modified = True
                    if socials.get("instagram") or socials.get("facebook"):
                        total_found += 1

                # Broadcast progress via SSE
                mgr._broadcast({
                    "type": "genius_progress",
                    "artist": str(artist_name),
                    "found": bool(socials),
                    "socials": socials or {},
                    "processed": total_processed,
                    "total_found": total_found,
                })

            if modified:
                try:
                    wb.save(str(item.output_path))
                    print(f"[genius-pass] Saved {item.output_path.name} ({total_found} socials found)", flush=True)
                except Exception as e:
                    print(f"[genius-pass] Save error: {e}", flush=True)

        print(f"[genius-pass] Complete. {total_processed} artists checked, {total_found} socials found.", flush=True)
        mgr._broadcast({"type": "genius_done", "processed": total_processed, "found": total_found})

    except Exception as e:
        print(f"[genius-pass] Error: {e}", flush=True)
    finally:
        _genius_running = False


# ---------------------------------------------------------------------------
# GENITRACTOR — Contact extraction tool (Genius-only, separate from main audit)
# Processes CSVs to extract Instagram/Facebook/YouTube via Genius API
# Exports a clean CSV with: Artist Name, Instagram, Facebook, YouTube
# ---------------------------------------------------------------------------

import csv
import io
import threading as _geni_threading
from queue import Queue as _GeniQueue, Full as _GeniFull

_geni_items = []
_geni_lock = _geni_threading.Lock()
_geni_subscribers = []
_geni_active_threads = {}
_geni_stop_flags = {}

GENI_UPLOAD_DIR = config.BASE_DIR / ".geni_uploads"
GENI_UPLOAD_DIR.mkdir(exist_ok=True)
GENI_OUTPUT_DIR = config.BASE_DIR / "GeniOutputs"
GENI_OUTPUT_DIR.mkdir(exist_ok=True)


@app.route("/api/cross-status")
def api_cross_status():
    """Returns status of both tools for cross-tool progress bar."""
    mgr = get_manager()

    # Chartporter status
    cp_processed = 0
    cp_total = 0
    cp_running = False
    with mgr._lock:
        for item in mgr._items:
            if item.status == "running":
                cp_running = True
                cp_processed += item.processed
                cp_total += item.total
            elif item.status == "done":
                cp_processed += item.processed
                cp_total += item.total

    # Genitractor status
    gn_processed = 0
    gn_total = 0
    gn_running = False
    with _geni_lock:
        for item in _geni_items:
            if item["status"] == "running":
                gn_running = True
            gn_processed += item.get("processed", 0)
            gn_total += item.get("total", 0)

    return jsonify({
        "chartporter": {"running": cp_running, "processed": cp_processed, "total": cp_total},
        "genitractor": {"running": gn_running, "processed": gn_processed, "total": gn_total},
    })


def _geni_broadcast(event):
    with _geni_lock:
        dead = []
        for q in _geni_subscribers:
            try:
                q.put_nowait(event)
            except _GeniFull:
                dead.append(q)
        for q in dead:
            try:
                _geni_subscribers.remove(q)
            except ValueError:
                pass


@app.route("/api/genitractor/upload", methods=["POST"])
def geni_upload():
    if "file" not in request.files:
        return jsonify({"error": "no file"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "no filename"}), 400
    ext = Path(f.filename).suffix.lower()
    if ext not in (".csv", ".tsv"):
        return jsonify({"error": "only .csv/.tsv"}), 400

    safe_name = f"{uuid.uuid4().hex[:8]}_{Path(f.filename).name}"
    dest = GENI_UPLOAD_DIR / safe_name
    f.save(str(dest))

    item = {
        "id": uuid.uuid4().hex[:12],
        "filename": f.filename,
        "path": str(dest),
        "status": "queued",
        "processed": 0,
        "total": 0,
        "error": "",
    }
    with _geni_lock:
        _geni_items.append(item)
    _geni_broadcast({"type": "item_added", "item": item})
    return jsonify(item), 201


@app.route("/api/genitractor/start", methods=["POST"])
def geni_start():
    with _geni_lock:
        queued = [i for i in _geni_items if i["status"] == "queued"]
        running = sum(1 for i in _geni_items if i["status"] == "running")
        available = 4 - running
        to_start = queued[:max(0, available)]

    for item in to_start:
        _geni_stop_flags[item["id"]] = False
        t = _geni_threading.Thread(target=_geni_worker, args=(item,), daemon=True)
        _geni_active_threads[item["id"]] = t
        t.start()

    return jsonify({"ok": True, "started": len(to_start)})


@app.route("/api/genitractor/stop", methods=["POST"])
def geni_stop():
    with _geni_lock:
        for item in _geni_items:
            if item["status"] == "running":
                _geni_stop_flags[item["id"]] = True
    return jsonify({"ok": True})


@app.route("/api/genitractor/export")
def geni_export():
    """Export all found contacts as a CSV."""
    with _geni_lock:
        all_contacts = []
        for item in _geni_items:
            contacts = list(item.get("_contacts", []))
            all_contacts.extend(contacts)

    if not all_contacts:
        return jsonify({"error": "no contacts found yet"}), 404

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Artist Name", "Instagram", "Facebook", "YouTube", "Website", "Email"])
    for c in all_contacts:
        writer.writerow([c.get("artist", ""), c.get("instagram", ""), c.get("facebook", ""), c.get("youtube", ""), c.get("website", ""), c.get("email", "")])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=Genitractor_Contacts.csv"},
    )


@app.route("/api/genitractor/stream")
def geni_stream():
    q = _GeniQueue(maxsize=200)
    with _geni_lock:
        _geni_subscribers.append(q)
    # Send snapshot
    snapshot = {"type": "snapshot", "items": [_geni_item_dict(i) for i in _geni_items]}
    try:
        q.put_nowait(snapshot)
    except _GeniFull:
        pass

    def generate():
        try:
            while True:
                try:
                    event = q.get(timeout=30)
                    yield f"data: {json.dumps(event)}\n\n"
                except Exception:
                    yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"
        except GeneratorExit:
            pass
        finally:
            with _geni_lock:
                try:
                    _geni_subscribers.remove(q)
                except ValueError:
                    pass

    return Response(generate(), mimetype="text/event-stream", headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


def _geni_item_dict(item):
    return {
        "id": item["id"],
        "filename": item["filename"],
        "status": item["status"],
        "processed": item["processed"],
        "total": item["total"],
        "error": item.get("error", ""),
    }


def _geni_worker(item):
    """Process one CSV — extract artist names, look up Genius socials one by one."""
    import time as _time
    import pandas as pd
    from app.sources import genius

    try:
        with _geni_lock:
            item["status"] = "running"
        _geni_broadcast({"type": "item_started", "item": _geni_item_dict(item)})

        # Read CSV
        df = pd.read_csv(item["path"])
        df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

        # Find artist column
        artist_col = None
        for col in df.columns:
            if col.lower().strip() in ("artist", "artist name", "name"):
                artist_col = col
                break
        if not artist_col:
            # Try second column if first is numeric ID
            if len(df.columns) >= 2:
                first_vals = df.iloc[:3, 0].astype(str)
                if all(v.isdigit() for v in first_vals):
                    artist_col = df.columns[1]

        if not artist_col:
            item["status"] = "error"
            item["error"] = "No artist column found"
            _geni_broadcast({"type": "item_error", "item": _geni_item_dict(item)})
            return

        artists = df[artist_col].dropna().astype(str).str.strip().tolist()
        item["total"] = len(artists)
        item["_contacts"] = []
        _geni_broadcast({"type": "item_started", "item": _geni_item_dict(item)})

        for i, artist_name in enumerate(artists):
            if _geni_stop_flags.get(item["id"]):
                item["status"] = "stopped"
                _geni_broadcast({"type": "item_stopped", "item": _geni_item_dict(item)})
                return

            if not artist_name:
                item["processed"] = i + 1
                continue

            # Rate limit: 1 request per 2 seconds
            _time.sleep(2.0)

            socials = genius.get_socials(artist_name)
            contact = {"artist": artist_name, "instagram": "", "facebook": "", "youtube": "", "website": "", "email": ""}

            if socials:
                if socials.get("instagram"):
                    contact["instagram"] = f"https://instagram.com/{socials['instagram']}"
                if socials.get("facebook"):
                    fb = socials["facebook"]
                    contact["facebook"] = fb if fb.startswith("http") else f"https://facebook.com/{fb}"
                if socials.get("youtube"):
                    contact["youtube"] = socials["youtube"]

            # --- Email scraping (free, no API key) ---
            # Waterfall: website → Facebook → YouTube description
            from app.sources import email_scraper

            website_url = None
            email_found = ""

            # Strategy 1: Try artist IG handle as domain
            ig_handle = socials.get("instagram") if socials else None
            if ig_handle:
                website_url = email_scraper.find_artist_website(ig_handle)

            # Strategy 2: Scrape emails from website if found
            if website_url:
                contact["website"] = website_url
                email_result = email_scraper.scrape_website_emails(website_url)
                if email_result and email_result.get("emails"):
                    email_found = email_result["emails"][0]

            # Strategy 3: Scrape Facebook page for email
            if not email_found and socials and socials.get("facebook"):
                fb_email = email_scraper.scrape_facebook_email(socials["facebook"])
                if fb_email:
                    email_found = fb_email

            # Strategy 4: Scrape YouTube description for email
            if not email_found:
                yt_email = email_scraper.scrape_youtube_description(artist_name)
                if yt_email:
                    email_found = yt_email

            if email_found:
                contact["email"] = email_found

            item["_contacts"].append(contact)
            item["processed"] = i + 1

            _geni_broadcast({
                "type": "contact_done",
                "item_id": item["id"],
                "artist": artist_name,
                "socials": socials,
                "website": contact["website"],
                "email": contact["email"],
                "processed": item["processed"],
                "total": item["total"],
            })

        item["status"] = "done"
        _geni_broadcast({"type": "item_done", "item": _geni_item_dict(item)})

        # Auto-start next queued
        with _geni_lock:
            next_queued = [i for i in _geni_items if i["status"] == "queued"][:1]
        for ni in next_queued:
            _geni_stop_flags[ni["id"]] = False
            t = _geni_threading.Thread(target=_geni_worker, args=(ni,), daemon=True)
            _geni_active_threads[ni["id"]] = t
            t.start()

    except Exception as e:
        import traceback
        print(f"[genitractor] Error: {e}\n{traceback.format_exc()}", flush=True)
        item["status"] = "error"
        item["error"] = str(e)
        _geni_broadcast({"type": "item_error", "item": _geni_item_dict(item)})
    finally:
        _geni_active_threads.pop(item["id"], None)
        # Cleanup upload file
        try:
            Path(item["path"]).unlink(missing_ok=True)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Feedback — writes .md files to feedback/ folder, Groq AI cleanup
# ---------------------------------------------------------------------------

FEEDBACK_DIR = config.BASE_DIR / "feedback"
FEEDBACK_DIR.mkdir(exist_ok=True)


@app.route("/api/feedback", methods=["POST"])
def api_feedback():
    """Submit feedback — writes a markdown file to feedback/ folder."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "invalid body"}), 400

    category = (data.get("category") or "").strip().upper()
    if category not in ("BUG", "IDEA", "OTHER"):
        return jsonify({"error": "category must be BUG, IDEA, or OTHER"}), 400

    text = (data.get("text") or "").strip()
    if not text:
        return jsonify({"error": "text is required"}), 400

    raw_text = (data.get("raw_text") or "").strip()
    ai_enhanced = bool(data.get("ai_enhanced", False))

    from datetime import datetime
    now = datetime.now()
    filename = f"{now.strftime('%m-%d.%H.%M')}.{category}.md"

    # Build markdown optimized for Claude parsing
    lines = [
        "---",
        f"category: {category}",
        f"date: {now.isoformat()}",
        f"platform: IGNITE VIRTUAL SCOUT",
        f"version: v5.0.0",
        f"ai_enhanced: {str(ai_enhanced).lower()}",
        "---",
        "",
        f"# {category}: {text.split(chr(10))[0][:80]}",
        "",
        text,
        "",
    ]

    if ai_enhanced and raw_text:
        lines.extend([
            "---",
            "",
            f"> **Original (raw):** {raw_text}",
            "",
        ])

    content = "\n".join(lines)
    filepath = FEEDBACK_DIR / filename
    filepath.write_text(content, encoding="utf-8")

    return jsonify({"ok": True, "file": filename}), 201


@app.route("/api/feedback/clean", methods=["POST"])
def api_feedback_clean():
    """Use Groq to clean up feedback text into an optimized Claude prompt."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "invalid body"}), 400

    text = (data.get("text") or "").strip()
    if not text:
        return jsonify({"error": "text is required"}), 400

    category = (data.get("category") or "OTHER").strip().upper()

    groq_key = config.groq_api_key()
    if not groq_key:
        return jsonify({"error": "Groq API key not configured. Add it in the API panel."}), 400

    category_context = {
        "BUG": "This is a bug report for IGNITE: VIRTUAL SCOUT, a Flask-based catalog intelligence platform. It processes CSV artist exports through iTunes, Deezer, Genius, Chartmetric, Groq, and Gemini APIs to verify catalog ownership for licensing/buyout opportunities.",
        "IDEA": "This is a feature idea for IGNITE: VIRTUAL SCOUT, a Flask-based catalog intelligence platform that processes CSV artist exports through multiple APIs to verify catalog ownership.",
        "OTHER": "This is general feedback for IGNITE: VIRTUAL SCOUT, a Flask-based catalog intelligence platform.",
    }.get(category, "This is feedback for IGNITE: VIRTUAL SCOUT.")

    system_prompt = f"""You are an expert prompt engineer. Your job is to take raw user feedback and transform it into a perfectly structured, actionable prompt that Claude (Opus) can immediately research and act on.

Rules:
- Fix all grammar and spelling errors
- Clarify vague instructions — ask yourself what Claude would need to know to act on this
- Add context about WHAT part of the system this relates to
- Break complex feedback into clear, actionable steps
- For bugs: include what happened, what should happen, and where it occurs
- For ideas: include the goal, how it would work, and what it affects
- Structure with markdown headers and bullet points
- Keep it concise but complete — no fluff
- Do NOT wrap in code fences or add explanatory text around the output
- Output ONLY the enhanced prompt text, nothing else

{category_context}"""

    import requests as http_req
    try:
        resp = http_req.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {groq_key}",
            },
            json={
                "model": "llama-3.3-70b-versatile",
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": text},
                ],
                "temperature": 0.3,
                "max_tokens": 1024,
            },
            timeout=30,
        )
        resp.raise_for_status()
        result = resp.json()
        cleaned = result["choices"][0]["message"]["content"].strip()
        return jsonify({"ok": True, "cleaned": cleaned})
    except http_req.exceptions.Timeout:
        return jsonify({"error": "Groq API timeout — try again"}), 504
    except http_req.exceptions.HTTPError as e:
        msg = str(e)
        try:
            msg = e.response.json().get("error", {}).get("message", msg)
        except Exception:
            pass
        return jsonify({"error": f"Groq API error: {msg}"}), 502
    except Exception as e:
        return jsonify({"error": f"Groq request failed: {e}"}), 500


# ---------------------------------------------------------------------------
# SSE stream — no Connection header (that was the Waitress crash)
# ---------------------------------------------------------------------------

@app.route("/api/stream")
def api_stream():
    """Server-Sent Events stream for live progress updates."""
    mgr = get_manager()
    q = mgr.subscribe()

    def generate():
        try:
            while True:
                try:
                    event = q.get(timeout=30)
                    yield f"data: {json.dumps(event)}\n\n"
                except Exception:
                    # Send heartbeat to keep connection alive
                    yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"
        except GeneratorExit:
            pass
        finally:
            mgr.unsubscribe(q)

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            # NOTE: Do NOT set 'Connection' header here.
            # Waitress (WSGI) rejects hop-by-hop headers from the app.
        },
    )
